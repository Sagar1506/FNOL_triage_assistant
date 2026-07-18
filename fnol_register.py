# -*- coding: utf-8 -*-
"""
fnol_register.py
================
FNOL registration module.

Functions:
  1. look_up_member()    - validate Member ID against policy_register_sample.xlsx
  2. generate_fnol_number() - get next FNOL number from fnol_register_sample.xlsx
  3. save_intake_document() - copy uploaded PDF to fnol_input folder
  4. register_fnol()     - write new row to fnol_register_sample.xlsx
  5. look_up_fnol()      - query existing FNOL by ID

All file paths are read from .env:
  POLICY_REGISTER_PATH  - path to policy_register_sample.xlsx
  FNOL_REGISTER_PATH    - path to fnol_register_sample.xlsx
  FNOL_INPUT_FOLDER     - folder where intake PDFs are saved
"""

import os
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

# Paths from .env
def _policy_register_path():
    return os.getenv("POLICY_REGISTER_PATH", "./Excel_data_base/policy_register_sample.xlsx")

def _fnol_register_path():
    return os.getenv("FNOL_REGISTER_PATH", "./Excel_data_base/fnol_register_sample.xlsx")

def _fnol_input_folder():
    return os.getenv("FNOL_INPUT_FOLDER", "./fnol/fnol_input")

# Styling helpers
_thin = Side(style="thin", color="CCCCCC")
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_al   = Alignment(wrap_text=True, horizontal="left", vertical="top")

def _style(cell, bg="FFFFFF", fg="3A3A3A", bold=False, size=10):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.border    = _bdr
    cell.alignment = _al


# 1. MEMBER LOOKUP
def look_up_member(member_id):
    """
    Look up a member_id in policy_register_sample.xlsx.
    Returns dict with member details if found, None otherwise.
    """
    path = _policy_register_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Policy register not found at: {path}\n"
            "Set POLICY_REGISTER_PATH in your .env file."
        )

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    member_policies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if str(d.get("member_id", "")).strip().upper() == member_id.strip().upper():
            member_policies.append(d)

    wb.close()

    if not member_policies:
        return None

    return {
        "member_id": member_id.strip().upper(),
        "policies":  member_policies,
    }


# 2. FNOL NUMBER GENERATION
def generate_fnol_number():
    """
    Returns the next FNOL number in the series.
    Format: FNOL-YYYY-NNNN (e.g. FNOL-2026-0003 -> FNOL-2026-0004)
    """
    path = _fnol_register_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"FNOL register not found at: {path}\n"
            "Set FNOL_REGISTER_PATH in your .env file."
        )

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    max_seq = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d     = dict(zip(headers, row))
        fid   = str(d.get("fnol_id") or "")
        parts = fid.split("-")
        if len(parts) == 3 and parts[0] == "FNOL":
            try:
                seq = int(parts[2])
                if seq > max_seq:
                    max_seq = seq
            except ValueError:
                pass

    wb.close()

    current_year = datetime.now().year
    new_seq      = max_seq + 1
    return f"FNOL-{current_year}-{new_seq:04d}"


# 3. SAVE INTAKE DOCUMENT
def save_intake_document(tmp_pdf_path, fnol_id):
    """
    Copies the uploaded PDF to:
        {FNOL_INPUT_FOLDER}/{fnol_id}/{fnol_id}_intake.pdf
    Returns the relative path stored in the register.
    """
    folder   = _fnol_input_folder()
    dest_dir = Path(folder) / fnol_id
    dest_dir.mkdir(parents=True, exist_ok=True)

    dest_file = dest_dir / f"{fnol_id}_intake.pdf"
    shutil.copy2(tmp_pdf_path, dest_file)

    return str(Path(folder).name / Path(fnol_id) / f"{fnol_id}_intake.pdf")


# 4. REGISTER FNOL
_FNOL_COLUMNS = [
    "fnol_id", "member_id", "policy_number", "vehicle_reg",
    "incident_date", "incident_type", "fnol_received_at",
    "fnol_age_hours", "aging_risk", "adjuster_id",
    "triage_status", "escalation_flag", "escalation_reason",
    "triage_json_path", "triage_pdf_path", "doc_links", "registered_at",
]

_COL_STYLES = {
    "fnol_id":         ("D5E8F0", "1F4E79", True),
    "member_id":       ("D5E8F0", "1F4E79", False),
    "policy_number":   ("D5E8F0", "1F4E79", False),
    "triage_status":   ("FFF2CC", "7B4F00", True),
    "escalation_flag": ("FCE4D6", "8B0000", True),
}


def register_fnol(fnol_id, member_id, policy_number, vehicle_reg,
                  incident_date, incident_type, adjuster_id, doc_links):
    """
    Appends a new FNOL row to fnol_register_sample.xlsx.
    Returns the fnol_id that was registered.
    """
    path = _fnol_register_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"FNOL register not found at: {path}\n"
            "Set FNOL_REGISTER_PATH in your .env file."
        )

    now_str = datetime.now().strftime("%d-%b-%Y %H:%M")

    new_row = {
        "fnol_id":           fnol_id,
        "member_id":         member_id,
        "policy_number":     policy_number,
        "vehicle_reg":       vehicle_reg,
        "incident_date":     incident_date,
        "incident_type":     incident_type,
        "fnol_received_at":  now_str,
        "fnol_age_hours":    0,
        "aging_risk":        "LOW",
        "adjuster_id":       adjuster_id,
        "triage_status":     "PENDING",
        "escalation_flag":   False,
        "escalation_reason": None,
        "triage_json_path":  None,
        "triage_pdf_path":   None,
        "doc_links":         doc_links,
        "registered_at":     now_str,
    }

    wb       = openpyxl.load_workbook(path)
    ws       = wb.active
    next_row = ws.max_row + 1
    print(f"[register_fnol] Writing '{fnol_id}' to row {next_row} of: {os.path.abspath(path)}")  # TEMP DIAGNOSTIC — remove once write location is confirmed
    bg_def   = "FFFFFF" if next_row % 2 == 0 else "F2F2F2"

    for ci, col in enumerate(_FNOL_COLUMNS, 1):
        val  = new_row.get(col)
        cell = ws.cell(row=next_row, column=ci, value=val)
        if col in _COL_STYLES:
            bg, fg, bold = _COL_STYLES[col]
            _style(cell, bg, fg, bold)
        else:
            _style(cell, bg_def)
        ws.row_dimensions[next_row].height = 30

    wb.save(path)
    wb.close()
    print(f"[register_fnol] Save complete. File modified at: "
          f"{datetime.fromtimestamp(os.path.getmtime(path)).strftime('%d-%b-%Y %H:%M:%S')}")  # TEMP DIAGNOSTIC — remove once write location is confirmed
    return fnol_id

# 5. FNOL LOOKUP
def look_up_fnol(fnol_id):
    """
    Look up a single FNOL record by fnol_id in fnol_register_sample.xlsx.
    Returns dict with all register columns if found, None otherwise.
    """
    path = _fnol_register_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"FNOL register not found at: {path}\n"
            "Set FNOL_REGISTER_PATH in your .env file."
        )

    wb      = openpyxl.load_workbook(path, read_only=True)
    ws      = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    result = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if str(d.get("fnol_id") or "").strip().upper() == fnol_id.strip().upper():
            result = d
            break

    wb.close()
    return result
