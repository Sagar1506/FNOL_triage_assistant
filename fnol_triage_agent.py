"""
fnol_triage_agent.py
====================
FNOL Triage Summary Generation Agent.

Reads FNOLIntakeState + Excel data files and produces:
  - Structured triage JSON (FNOLTriageState)
  - Updates fnol_register_sample.xlsx with triage_status, paths, escalation

Public API:
    run_triage_agent(intake_state, fnol_id) -> FNOLTriageState

LLM used for:
  - Coverage applicability analysis (Section 2)
  - Exclusion analysis (Section 3)
  All other sections are deterministic Python.

LLM model is configurable via LLM_MODEL in .env (default: gpt-4o-mini).
"""

import os
import json
import time
import logging
from datetime import datetime, date
from typing import TypedDict, Optional
from dateutil.parser import parse as parse_date

import openpyxl
from openai import OpenAI
from dotenv import load_dotenv

# Single source of truth for the FNOL register path — previously this file
# defined its own copy of this helper with a DIFFERENT default
# ("./fnol_register_sample.xlsx" vs fnol_register.py's
# "./Excel_data_base/fnol_register_sample.xlsx"). When FNOL_REGISTER_PATH
# wasn't set in .env, register_fnol() (called from main.py) and
# update_fnol_register() (called below) silently wrote to two different
# files — new rows appeared to "not save" because the update step was
# reading/writing a file that never had the row in the first place.
from fnol_register import _fnol_register_path

load_dotenv()

logger = logging.getLogger(__name__)

# ── LangSmith tracing (optional — only active when LANGCHAIN_TRACING_V2=true)
try:
    from langsmith import traceable
except ImportError:
    def traceable(*args, **kwargs):
        def decorator(fn): return fn
        return decorator if args and callable(args[0]) else decorator

# ── LLM client — lazy initialisation so .env is loaded before client created
LLM_MODEL = os.getenv("LLM_MODEL", "gpt-4o-mini")
_client   = None

def _get_client():
    """Return OpenAI client, initialising lazily after .env is loaded."""
    global _client
    if _client is None:
        _client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    return _client


# ── Paths from .env ────────────────────────────────────────────────────────
def _policy_register_path():
    return os.getenv("POLICY_REGISTER_PATH", "./Excel_data_base/policy_register_sample.xlsx")

def _policy_clauses_path():
    return os.getenv("POLICY_CLAUSES_PATH", "./Excel_data_base/policy_clauses_sample.xlsx")

def _policy_exclusions_path():
    return os.getenv("POLICY_EXCLUSIONS_PATH", "./Excel_data_base/policy_exclusions_sample.xlsx")

def _claims_history_path():
    return os.getenv("CLAIMS_HISTORY_PATH", "./Excel_data_base/claims_history_sample.xlsx")

def _triage_output_path():
    return os.getenv("TRIAGE_OUTPUT_PATH", "./fnol/triage_output")


# ══════════════════════════════════════════════════════════════════════════
# STATE
# ══════════════════════════════════════════════════════════════════════════

class FNOLTriageState(TypedDict):
    # Identity
    fnol_id:               str
    triage_generated_at:   str
    adjuster_id:           str

    # Sections
    claim_snapshot:        dict
    aging_risk:            dict
    coverage_result:       list[dict]
    coverage_applicable:   bool
    coverage_confidence_overall: str
    exclusion_result:      list[dict]
    waiting_period:        dict
    claims_history:        dict
    fraud_signals:         dict
    doc_checklist:         list[dict]
    escalation:            dict

    # Final status
    triage_status:         str   # COMPLETE|ESCALATED|PENDED_FOR_INPUTS|EXCEPTION_RAISED
    escalation_flag:       bool
    escalation_reason:     Optional[str]

    # Output paths
    triage_json_path:      Optional[str]
    triage_pdf_path:       Optional[str]

    # timing_report: real, measured per-step durations for THIS run —
    # same pattern as fnol_intake_agent.py's timing_report. {"total_duration_s":
    # float, "steps": [{"step", "duration_s", "start_offset_s"}, ...]}
    timing_report:          dict
    # timing_report_path: where timing_report was saved as JSON, alongside
    # the triage JSON/PDF for this fnol_id.
    timing_report_path:     Optional[str]

    # Agent metadata
    agent_status:          str
    error_message:         Optional[str]


# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════

def _read_excel_rows(path: str) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows    = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return headers, rows


def _safe_date(val) -> Optional[date]:
    """Parse a date value from various formats. Returns None on failure."""
    if val is None:
        return None
    try:
        if isinstance(val, (datetime, date)):
            return val if isinstance(val, date) else val.date()
        return parse_date(str(val), dayfirst=True).date()
    except Exception:
        return None


@traceable(name="triage._call_llm_json", run_type="llm")
def _call_llm_json(system: str, user: str) -> dict:
    resp = _get_client().chat.completions.create(
        model=LLM_MODEL,
        temperature=0.0,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
    )
    return json.loads(resp.choices[0].message.content.strip())


# ══════════════════════════════════════════════════════════════════════════
# SECTION 1 — CLAIM SNAPSHOT
# Pure Python — policy register lookup
# ══════════════════════════════════════════════════════════════════════════

def build_claim_snapshot(fnol_fields: dict, fnol_id: str,
                          fnol_received_at: str) -> tuple[dict, dict]:
    """
    Looks up policy from policy_register_sample.xlsx.
    Returns (claim_snapshot, policy_row).
    Raises ValueError if policy not found.
    """
    policy_number = fnol_fields.get("policy_number", "")
    vehicle_reg   = fnol_fields.get("vehicle_registration_number", "")

    _, rows = _read_excel_rows(_policy_register_path())

    # Match by policy_number first, fallback to vehicle_reg
    policy = None
    for r in rows:
        if str(r.get("policy_number", "")).strip() == policy_number.strip():
            policy = r
            break
    if not policy:
        for r in rows:
            if str(r.get("vehicle_reg", "")).strip().upper() == vehicle_reg.strip().upper():
                policy = r
                break

    if not policy:
        raise ValueError(f"No policy found for policy_number={policy_number} "
                         f"or vehicle_reg={vehicle_reg}")

    # Policy expiry warning
    incident_date  = _safe_date(fnol_fields.get("incident_date_time"))
    policy_expiry  = _safe_date(policy.get("policy_expiry"))
    expiry_warning = None
    if incident_date and policy_expiry and incident_date > policy_expiry:
        expiry_warning = (
            f"Policy {policy['policy_number']} expired "
            f"{policy['policy_expiry']}. Incident date "
            f"{fnol_fields.get('incident_date_time', '')} is after expiry. "
            f"Coverage requires renewal verification."
        )

    snapshot = {
        # Prefer the policy register's member_id (verified, tied to this
        # exact policy_number/vehicle_reg match) over the FNOL-extracted
        # value, which can be missing — see check_claims_history() for the
        # same fix and full rationale.
        "member_id":        policy.get("member_id") or fnol_fields.get("member_id"),
        "policy_number":    policy["policy_number"],
        "vehicle_reg":      policy["vehicle_reg"],
        "vehicle":          f"{policy['vehicle_make']} {policy['vehicle_model']} ({policy['vehicle_year']})",
        "incident_type":    fnol_fields.get("incident_type"),
        "incident_date":    fnol_fields.get("incident_date_time"),
        "incident_location":fnol_fields.get("incident_location"),
        "fnol_received_at": fnol_received_at,
        "policy_status":    policy.get("policy_status"),
        "policy_inception": policy.get("policy_inception"),
        "policy_expiry":    policy.get("policy_expiry"),
        "coverage_type":    policy.get("coverage_type"),
        "idv_inr":          policy.get("idv"),
        "ncb_percent":      policy.get("ncb_percent"),
        "addons":           str(policy.get("addons") or "").split("|") if policy.get("addons") else [],
        "hypothecated":     bool(policy.get("hypothecated")),
        "financer_name":    policy.get("financer_name"),
        "insurer":          policy.get("insurer_name"),
        "policy_expiry_warning": expiry_warning,
        "waiting_period_days":   int(policy.get("waiting_period_days") or 30),
    }
    return snapshot, policy


# ══════════════════════════════════════════════════════════════════════════
# SECTION 2 — COVERAGE APPLICABILITY
# Deterministic pre-checks + LLM for clause interpretation
# ══════════════════════════════════════════════════════════════════════════

COVERAGE_ANALYSIS_SYSTEM = f"""
You are a motor insurance coverage analysis expert for Indian motor insurance.
You will receive a FNOL incident description and a list of policy clauses.
For each clause, determine coverage applicability.

Return a JSON object with key "coverage_result" containing a list.
Each item in the list must have:
  "section"       : the section_name from the clause
  "clause_ref"    : the clause_ref from the clause
  "status"        : one of COVERED | NOT_COVERED | REQUIRES_VERIFICATION | NOT_APPLICABLE
  "confidence"    : HIGH | MODERATE | LOW
  "note"          : one concise sentence explaining the determination

Rules:
- COVERED: clause clearly applies, no blockers
- NOT_COVERED: clause clearly does not apply to this incident type or policy type
- REQUIRES_VERIFICATION: data missing, document absent, or ambiguity present
- NOT_APPLICABLE: add-on clause for an add-on not in the policy, or incident type mismatch
- If policy is expired and incident is after expiry, set status=REQUIRES_VERIFICATION,
  confidence=LOW, note must mention the expiry issue
- Never use the word "fraud" or make fraud allegations
- Model: {LLM_MODEL}
"""


@traceable(name="triage.analyse_coverage", run_type="chain")
def analyse_coverage(fnol_fields: dict, policy_row: dict,
                     claim_snapshot: dict) -> tuple[list[dict], bool, str]:
    """
    Returns (coverage_result, coverage_applicable, confidence_overall).
    Step 1: deterministic pre-checks (policy status, expiry, waiting period)
    Step 2: LLM clause interpretation
    """
    incident_type  = str(fnol_fields.get("incident_type", "")).strip().upper().replace(" ", "_")
    policy_number  = policy_row["policy_number"]
    coverage_type  = policy_row.get("coverage_type", "")
    addons         = str(policy_row.get("addons") or "").split("|")
    policy_status  = str(policy_row.get("policy_status", "")).upper()

    # ── Deterministic pre-check: lapsed/cancelled ─────────────────────────
    if policy_status in ("LAPSED", "CANCELLED"):
        return ([{
            "section":    "All Sections",
            "clause_ref": "N/A",
            "status":     "NOT_COVERED",
            "confidence": "HIGH",
            "note":       f"Policy is {policy_status}. No coverage available."
        }], False, "HIGH")

    # ── Load relevant clauses ─────────────────────────────────────────────
    _, clause_rows = _read_excel_rows(_policy_clauses_path())

    relevant = []
    for cl in clause_rows:
        if cl.get("policy_number") != policy_number:
            continue
        applicable_types = str(cl.get("applicable_incident_types") or "").split("|")
        cov_filter       = str(cl.get("coverage_type_filter") or "")
        addon_req        = str(cl.get("addon_required") or "").strip()

        # Filter by incident type
        if incident_type not in applicable_types and "ALL" not in applicable_types:
            continue
        # Filter by coverage type
        if cov_filter and coverage_type not in cov_filter.split("|"):
            continue
        # Skip add-on clauses if add-on not in policy
        if addon_req and addon_req not in addons:
            continue
        relevant.append(cl)

    if not relevant:
        return ([{
            "section":    "No applicable clauses",
            "clause_ref": "N/A",
            "status":     "REQUIRES_VERIFICATION",
            "confidence": "LOW",
            "note":       "No matching clauses found in policy_clauses for this incident type."
        }], False, "LOW")

    # ── Build LLM input ───────────────────────────────────────────────────
    clauses_text = "\n\n".join([
        f"CLAUSE {i+1}:\n"
        f"  section_name: {cl['section_name']}\n"
        f"  clause_ref: {cl['clause_ref']}\n"
        f"  clause_text: {cl['clause_text']}\n"
        f"  coverage_status_default: {cl['coverage_status_default']}\n"
        f"  addon_required: {cl.get('addon_required') or 'None'}"
        for i, cl in enumerate(relevant)
    ])

    user_prompt = f"""FNOL DETAILS:
  Incident type     : {incident_type}
  Incident date     : {fnol_fields.get('incident_date_time')}
  Damage description: {fnol_fields.get('damage_description')}
  Third party       : {fnol_fields.get('third_party_involved')}
  TP injury         : {fnol_fields.get('third_party_injury_reported')}
  FIR filed         : {fnol_fields.get('fir_filed')}
  Legal notice      : {fnol_fields.get('legal_notice_present')}

POLICY DETAILS:
  Policy status     : {policy_status}
  Coverage type     : {coverage_type}
  IDV               : Rs.{policy_row.get('idv', 0):,}
  Add-ons active    : {', '.join(addons) if addons else 'None'}
  Policy inception  : {policy_row.get('policy_inception')}
  Policy expiry     : {policy_row.get('policy_expiry')}
  Hypothecated      : {policy_row.get('hypothecated')}
  Financer          : {policy_row.get('financer_name') or 'None'}

POLICY EXPIRY WARNING: {claim_snapshot.get('policy_expiry_warning') or 'None'}

CLAUSES TO ANALYSE:
{clauses_text}

Analyse coverage applicability for each clause above."""

    result = _call_llm_json(COVERAGE_ANALYSIS_SYSTEM, user_prompt)
    coverage_list = result.get("coverage_result", [])

    # Determine overall confidence
    statuses = [c.get("confidence", "LOW") for c in coverage_list]
    if all(s == "HIGH" for s in statuses):
        overall = "HIGH"
    elif any(s == "LOW" for s in statuses):
        overall = "LOW"
    else:
        overall = "MODERATE"

    applicable = any(
        c.get("status") == "COVERED"
        for c in coverage_list
    )

    return coverage_list, applicable, overall


# ══════════════════════════════════════════════════════════════════════════
# SECTION 3 — EXCLUSION ANALYSIS
# LLM interprets each exclusion against incident facts
# ══════════════════════════════════════════════════════════════════════════

EXCLUSION_ANALYSIS_SYSTEM = f"""
You are a motor insurance exclusion analysis expert for Indian motor insurance.
You will receive FNOL incident details and a list of policy exclusions.
For each exclusion, determine whether it may be relevant to this claim.

Return a JSON object with key "exclusion_result" containing a list.
Each item must have:
  "code"            : the exclusion_code
  "name"            : the exclusion_name
  "clause_ref"      : the clause_ref
  "applicability"   : REQUIRES_VERIFICATION (always — adjuster decides)
  "confidence"      : HIGH | MODERATE | LOW
  "reason_relevant" : one concise sentence explaining why this exclusion
                      is or is not relevant to this specific incident

Rules:
- Never declare an exclusion as definitively "Confirmed" — always REQUIRES_VERIFICATION
- If the exclusion is clearly NOT relevant (e.g. WAR_NUCLEAR for a road accident), 
  set confidence=LOW and state it does not appear to apply
- If the exclusion IS potentially relevant, set confidence=HIGH or MODERATE
- Never use the word "fraud" or make fraud allegations
- Model: {LLM_MODEL}
"""


@traceable(name="triage.analyse_exclusions", run_type="chain")
def analyse_exclusions(fnol_fields: dict, policy_row: dict) -> list[dict]:
    """Returns list of exclusion analysis results."""
    incident_type = str(fnol_fields.get("incident_type", "")).strip().upper().replace(" ", "_")
    policy_number = policy_row["policy_number"]
    addons        = str(policy_row.get("addons") or "").split("|")

    _, excl_rows = _read_excel_rows(_policy_exclusions_path())

    relevant = []
    for ex in excl_rows:
        if ex.get("policy_number") != policy_number:
            continue
        if not ex.get("active", True):
            continue
        applicable = str(ex.get("applicable_incident_types") or "ALL")
        if incident_type not in applicable.split("|") and "ALL" not in applicable:
            continue
        # Skip waivable exclusions if reinstatement add-on is present
        reinstatement = str(ex.get("reinstatement_addon") or "").strip()
        if reinstatement and reinstatement in addons:
            continue
        relevant.append(ex)

    # Sort by display_order
    relevant.sort(key=lambda x: int(x.get("display_order") or 99))

    if not relevant:
        return []

    exclusions_text = "\n\n".join([
        f"EXCLUSION {i+1}:\n"
        f"  exclusion_code: {ex['exclusion_code']}\n"
        f"  exclusion_name: {ex['exclusion_name']}\n"
        f"  clause_ref: {ex['clause_ref']}\n"
        f"  clause_text: {ex['clause_text']}\n"
        f"  trigger_field: {ex.get('trigger_field') or 'N/A'}\n"
        f"  trigger_condition: {ex.get('trigger_condition') or 'N/A'}\n"
        f"  confidence_if_triggered: {ex.get('confidence_if_triggered') or 'HIGH'}"
        for i, ex in enumerate(relevant)
    ])

    user_prompt = f"""FNOL DETAILS:
  Incident type      : {incident_type}
  Damage description : {fnol_fields.get('damage_description')}
  Third party        : {fnol_fields.get('third_party_involved')}
  TP injury          : {fnol_fields.get('third_party_injury_reported')}
  FIR filed          : {fnol_fields.get('fir_filed')}
  DL submitted       : {'Yes' if fnol_fields.get('fir_filed') else 'Unknown'}
  Legal notice       : {fnol_fields.get('legal_notice_present')}

POLICY ADD-ONS ACTIVE: {', '.join(addons) if addons else 'None'}

EXCLUSIONS TO ANALYSE:
{exclusions_text}

Analyse each exclusion for relevance to this specific incident."""

    result = _call_llm_json(EXCLUSION_ANALYSIS_SYSTEM, user_prompt)
    return result.get("exclusion_result", [])


# ══════════════════════════════════════════════════════════════════════════
# SECTION 4 — WAITING PERIOD
# Pure Python — date arithmetic
# ══════════════════════════════════════════════════════════════════════════

def check_waiting_period(fnol_fields: dict, policy_row: dict) -> dict:
    inception_str     = str(policy_row.get("policy_inception") or "")
    incident_str      = str(fnol_fields.get("incident_date_time") or "")
    waiting_days      = int(policy_row.get("waiting_period_days") or 30)

    inception_date    = _safe_date(inception_str)
    incident_date     = _safe_date(incident_str)

    if not inception_date or not incident_date:
        return {
            "policy_inception":       inception_str,
            "waiting_period_days":    waiting_days,
            "incident_date":          incident_str,
            "days_since_inception":   None,
            "waiting_period_active":  False,
            "incident_pre_inception": False,
            "confidence":             "LOW",
            "note":                   "Could not parse dates for waiting period calculation.",
        }

    days_since = (incident_date - inception_date).days
    waiting_active   = 0 <= days_since < waiting_days
    pre_inception    = incident_date < inception_date

    note = ""
    if pre_inception:
        note = "Incident date is before policy inception. Critical exception — immediate hold."
    elif waiting_active:
        note = (f"WAITING PERIOD ACTIVE — incident occurred {days_since} days after "
                f"inception. Mandatory waiting period is {waiting_days} days.")
    else:
        note = (f"Waiting period has elapsed. Policy has been in force for "
                f"{days_since} days at the incident date.")

    return {
        "policy_inception":       inception_str,
        "waiting_period_days":    waiting_days,
        "incident_date":          incident_str,
        "days_since_inception":   days_since,
        "waiting_period_active":  waiting_active,
        "incident_pre_inception": pre_inception,
        "confidence":             "HIGH",
        "note":                   note,
    }


# ══════════════════════════════════════════════════════════════════════════
# SECTION 5 — CLAIMS HISTORY
# Pure Python — Excel filter and count
# ══════════════════════════════════════════════════════════════════════════

def check_claims_history(fnol_fields: dict, policy_row: dict) -> dict:
    vehicle_reg   = str(policy_row.get("vehicle_reg") or "").strip().upper()
    # Prefer the member_id from the matched policy_register row — it's the
    # verified value tied to this policy_number/vehicle_reg lookup, and is
    # always present in policy_register_sample.xlsx. fnol_fields.get() is
    # only a fallback: it comes from re-extracting the uploaded FNOL PDF,
    # which can come back empty (e.g. PII-redaction routing, extraction
    # miss) and would otherwise silently zero out REPEAT_CLAIMANT for
    # every affected claim by comparing against "".
    member_id     = str(policy_row.get("member_id") or fnol_fields.get("member_id") or "").strip().upper()
    incident_date = _safe_date(fnol_fields.get("incident_date_time"))
    ncb_percent   = float(policy_row.get("ncb_percent") or 0)

    if not incident_date:
        return {"history_available": False,
                "note": "Could not parse incident date for history lookup."}

    _, history_rows = _read_excel_rows(_claims_history_path())

    # 36-month filter — same vehicle
    cutoff_36m = date(incident_date.year - 3, incident_date.month, incident_date.day)
    prior_36m  = []
    for r in history_rows:
        if str(r.get("vehicle_reg") or "").strip().upper() != vehicle_reg:
            continue
        claim_date = _safe_date(r.get("claim_date"))
        if claim_date and cutoff_36m <= claim_date < incident_date:
            prior_36m.append(r)

    # 12-month filter — same member, all vehicles
    cutoff_12m = date(incident_date.year - 1, incident_date.month, incident_date.day)
    prior_12m_count = sum(
        1 for r in history_rows
        if str(r.get("member_id") or "").strip().upper() == member_id
        and (lambda d: d and cutoff_12m <= d < incident_date)(_safe_date(r.get("claim_date")))
    )

    last_claim_date   = max(
        (_safe_date(r["claim_date"]) for r in prior_36m
         if _safe_date(r["claim_date"])),
        default=None
    )
    days_since_last   = (incident_date - last_claim_date).days if last_claim_date else None
    back_to_back      = days_since_last is not None and days_since_last <= 30
    repeat_claimant   = prior_12m_count >= 3

    # NCB discrepancy check
    ncb_discrepancy = None
    if ncb_percent > 0:
        ncb_impact_claims = [r for r in prior_36m if r.get("ncb_impact") is True]
        if ncb_impact_claims:
            ncb_discrepancy = (
                f"NCB is {ncb_percent:.0f}% on policy but "
                f"{len(ncb_impact_claims)} recent claim(s) show ncb_impact = TRUE. "
                f"Verify whether NCB was protected via NCB_PROTECT add-on."
            )

    return {
        "history_available":      True,
        "prior_claims_36m":       [
            {
                "claim_id":      r.get("claim_id"),
                "claim_date":    str(r.get("claim_date")),
                "incident_type": r.get("incident_type"),
                "status":        r.get("claim_status"),
                "settled_amount_inr": r.get("settled_amount"),
                "ncb_impact":    r.get("ncb_impact"),
            }
            for r in prior_36m
        ],
        "prior_claims_36m_count": len(prior_36m),
        "last_claim_date":        str(last_claim_date) if last_claim_date else None,
        "days_since_last_claim":  days_since_last,
        "back_to_back_flag":      back_to_back,
        "prior_claims_12m_count": prior_12m_count,
        "repeat_claimant_flag":   repeat_claimant,
        "ncb_discrepancy":        ncb_discrepancy,
    }


# ══════════════════════════════════════════════════════════════════════════
# SECTION 6 — DOCUMENT CHECKLIST
# Reused directly from intake state — no recalculation needed
# ══════════════════════════════════════════════════════════════════════════

def build_doc_checklist(intake_state: dict) -> tuple[list[dict], bool]:
    """
    Converts document_gap_checks from intake state into the triage
    doc_checklist format matching the sample JSON.
    """
    checklist = []
    critical_missing = False

    for gap in intake_state.get("document_gap_checks", []):
        if gap.get("status") == "SKIPPED":
            continue
        for doc in gap.get("documents", []):
            status   = doc["status"]
            blocking = doc["criticality"] in ("BLOCKING", "MANDATORY")
            received = status == "RECEIVED"

            if status == "MISSING_BLOCKING":
                critical_missing = True

            checklist.append({
                "name":             doc["document"],
                "mandatory":        blocking,
                "blocking":         doc["criticality"] == "BLOCKING",
                "received":         received,
                "status":           "RECEIVED" if received else "MISSING",
                "collection_target":doc.get("collection_by"),
            })

    return checklist, critical_missing


# ══════════════════════════════════════════════════════════════════════════
# SECTION 7 — ESCALATION
# Pure Python — 10 deterministic conditions
# ══════════════════════════════════════════════════════════════════════════

def evaluate_escalation(fnol_fields: dict, policy_row: dict,
                         waiting_period: dict, claims_history: dict,
                         fraud_signals: dict,
                         fnol_id: str, fnol_received_at: str) -> dict:
    """
    Evaluates all 10 escalation conditions, plus one additional condition
    (ELEVATED_REVIEW_INDICATORS) driven by the fraud signal agent's HIGH
    band. Returns escalation dict with triggered_conditions list.
    """
    triggered     = []
    incident_type = str(fnol_fields.get("incident_type", "")).strip().upper().replace(" ", "_")
    idv           = float(policy_row.get("idv") or 0)
    hypothecated  = bool(policy_row.get("hypothecated"))
    policy_status = str(policy_row.get("policy_status", "")).upper()
    inception_str = str(policy_row.get("policy_inception") or "")
    incident_date = _safe_date(fnol_fields.get("incident_date_time"))
    fnol_received = _safe_date(fnol_received_at)

    def add(code, label, trigger, instruction, **extra):
        entry = {"code": code, "label": label,
                 "trigger": trigger, "adjuster_instruction": instruction}
        entry.update(extra)
        triggered.append(entry)

    # 1. HIGH_VALUE_CLAIM
    high_value_types = {"THEFT_COMPLETE", "ACCIDENT_WITH_THIRD_PARTY_INJURY", "FIRE"}
    if idv > 500000 and incident_type in high_value_types:
        add("HIGH_VALUE_CLAIM",
            "High-Value Claim",
            f"IDV Rs.{idv:,.0f} > Rs.5,00,000 AND incident type is {incident_type}",
            "Refer to senior claims manager before proceeding.")

    # 2. THIRD_PARTY_BODILY_INJURY
    tp_injury = str(fnol_fields.get("third_party_injury_reported") or "").lower()
    if tp_injury.startswith("yes") or tp_injury == "true":
        add("THIRD_PARTY_BODILY_INJURY",
            "Third-Party Bodily Injury Reported",
            "tp_injury_reported = TRUE on FNOL form",
            "Refer to senior claims manager. Do not make any liability statement to claimant. Legal exposure.")

    # 3. THEFT_WITH_ACTIVE_LOAN
    if incident_type == "THEFT_COMPLETE" and hypothecated:
        add("THEFT_WITH_ACTIVE_LOAN",
            "Theft with Active Loan (Hypothecation)",
            "incident_type = THEFT_COMPLETE AND hypothecated = TRUE",
            "Financer notification required. NOC must be obtained. Refer to senior.")

    # 4. INCIDENT_IN_WAITING_PERIOD
    if waiting_period.get("waiting_period_active"):
        days = waiting_period.get("days_since_inception", "?")
        add("INCIDENT_IN_WAITING_PERIOD",
            "Incident in Mandatory Waiting Period",
            f"days_since_inception = {days} < waiting_period_days = {waiting_period.get('waiting_period_days')}",
            "Do not communicate any coverage finding. Senior review mandatory.")

    # 5. BACK_TO_BACK_CLAIMS
    days_since = claims_history.get("days_since_last_claim")
    if days_since is not None and days_since <= 30:
        prior_id = (claims_history.get("prior_claims_36m") or [{}])[0].get("claim_id", "")
        add("BACK_TO_BACK_CLAIMS",
            "Back-to-Back Claims",
            f"days_since_last_claim = {days_since} (threshold: 30 or fewer days)",
            "Senior review required before proceeding.",
            prior_claim_id=prior_id,
            prior_claim_date=claims_history.get("last_claim_date"))

    # 6. LATE_FNOL_OVER_7_DAYS
    if incident_date and fnol_received:
        days_late = (fnol_received - incident_date).days
        if days_late > 7:
            add("LATE_FNOL_OVER_7_DAYS",
                "Late FNOL — Over 7 Days",
                f"FNOL received {days_late} days after incident date",
                "Request written explanation from claimant. Refer to senior.")

    # 7. LEGAL_NOTICE_OR_TRIBUNAL
    legal = str(fnol_fields.get("legal_notice_present") or "").lower()
    if legal.startswith("yes") or legal == "true":
        add("LEGAL_NOTICE_OR_TRIBUNAL",
            "Legal Notice or Tribunal Reference",
            "legal_notice_present = TRUE on FNOL form",
            "Immediate hold. Do not contact claimant without legal team approval.")

    # 8. INCIDENT_BEFORE_POLICY_START
    if waiting_period.get("incident_pre_inception"):
        add("INCIDENT_BEFORE_POLICY_START",
            "Incident Before Policy Start Date",
            "incident_date < policy_inception_date",
            "Critical exception. Immediate hold. Do not communicate anything.")

    # 9. POLICY_NOT_FOUND — handled upstream; if policy_status is lapsed/cancelled
    if policy_status in ("LAPSED", "CANCELLED"):
        add("POLICY_NOT_FOUND",
            f"Policy Status: {policy_status}",
            f"policy_status = {policy_status}",
            "Immediate hold. All coverage sections unavailable. Verify with underwriting.")

    # 10. REPEAT_CLAIMANT
    if claims_history.get("repeat_claimant_flag"):
        count = claims_history.get("prior_claims_12m_count", 0)
        add("REPEAT_CLAIMANT",
            "Repeat Claimant",
            f"prior_claims_12m = {count} (threshold: 3 or more)",
            "Surface to senior for awareness. Does not block triage independently.")

    # 11. ELEVATED_REVIEW_INDICATORS — driven by the fraud signal agent.
    # Only a HIGH band escalates on its own; MEDIUM/LOW never do. Wording
    # deliberately avoids the word "fraud" per FNOL-GUIDE-001 Section 13.5.
    fraud_band = (fraud_signals or {}).get("band")
    if fraud_band == "HIGH":
        score = fraud_signals.get("fraud_score", 0)
        add("ELEVATED_REVIEW_INDICATORS",
            "Elevated Review Indicators",
            f"fraud_signals.band = HIGH (score = {score})",
            "Refer to SIU for review. Does not by itself constitute a finding.")

    escalation_flag   = len(triggered) > 0
    escalation_reason = "|".join(t["code"] for t in triggered) if triggered else None

    return {
        "escalation_flag":       escalation_flag,
        "triggered_conditions":  triggered,
        "escalation_reason":     escalation_reason,
    }


# ══════════════════════════════════════════════════════════════════════════
# SECTION 8 — AGING RISK
# Pure Python — time arithmetic
# ══════════════════════════════════════════════════════════════════════════

def compute_aging_risk(fnol_received_at: str) -> dict:
    now           = datetime.now()
    received      = None
    try:
        received  = parse_date(fnol_received_at, dayfirst=True)
    except Exception:
        pass

    if not received:
        return {"fnol_age_hours": 0, "aging_risk_level": "UNKNOWN",
                "sla_breached": False}

    age_hours = (now - received).total_seconds() / 3600
    if age_hours < 6:
        level = "LOW"
    elif age_hours < 18:
        level = "MEDIUM"
    elif age_hours < 24:
        level = "HIGH"
    else:
        level = "CRITICAL"

    sla_due = datetime(received.year, received.month, received.day,
                       received.hour, received.minute) 
    from datetime import timedelta
    sla_due = received + timedelta(hours=24)

    return {
        "fnol_age_hours":    round(age_hours, 2),
        "aging_risk_level":  level,
        "irdai_sla_due":     sla_due.strftime("%d-%b-%Y %H:%M"),
        "sla_breached":      age_hours > 24,
    }


# ══════════════════════════════════════════════════════════════════════════
# TRIAGE STATUS DETERMINATION
# ══════════════════════════════════════════════════════════════════════════

def determine_triage_status(policy_row: dict, escalation: dict,
                              critical_docs_missing: bool,
                              waiting_period: dict) -> str:
    """
    Priority order:
      1. EXCEPTION_RAISED — lapsed/cancelled policy
      2. ESCALATED        — any escalation condition triggered
      3. PENDED_FOR_INPUTS — blocking documents missing
      4. COMPLETE         — all clear
    """
    policy_status = str(policy_row.get("policy_status", "")).upper()
    if policy_status in ("LAPSED", "CANCELLED"):
        return "EXCEPTION_RAISED"
    if escalation.get("escalation_flag"):
        return "ESCALATED"
    if critical_docs_missing:
        return "PENDED_FOR_INPUTS"
    return "COMPLETE"


# ══════════════════════════════════════════════════════════════════════════
# FNOL REGISTER UPDATE
# ══════════════════════════════════════════════════════════════════════════

def update_fnol_register(fnol_id: str, triage_status: str,
                          escalation_flag: bool, escalation_reason: str,
                          triage_json_path: str, triage_pdf_path: str):
    """Update the FNOL register row with triage results."""
    path = _fnol_register_path()
    wb   = openpyxl.load_workbook(path)
    ws   = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    col_map = {h: i+1 for i, h in enumerate(headers)}

    row_found = False
    for row_num in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=col_map.get("fnol_id", 1)).value
        if str(cell_val or "").strip() == fnol_id.strip():
            row_found = True
            if "triage_status" in col_map:
                ws.cell(row=row_num, column=col_map["triage_status"]).value = triage_status
            if "escalation_flag" in col_map:
                ws.cell(row=row_num, column=col_map["escalation_flag"]).value = escalation_flag
            if "escalation_reason" in col_map:
                ws.cell(row=row_num, column=col_map["escalation_reason"]).value = escalation_reason
            if "triage_json_path" in col_map:
                ws.cell(row=row_num, column=col_map["triage_json_path"]).value = triage_json_path
            if "triage_pdf_path" in col_map:
                ws.cell(row=row_num, column=col_map["triage_pdf_path"]).value = triage_pdf_path
            break

    wb.save(path)
    wb.close()

    if not row_found:
        # Previously silent: saved the file unchanged and returned as if
        # nothing was wrong. That masked exactly the kind of path-mismatch
        # bug fixed above — if this ever fires again (wrong fnol_id passed
        # in, register file reset, etc.) it should be visible immediately
        # rather than surfacing later as "my records aren't saving."
        raise ValueError(
            f"update_fnol_register: no row with fnol_id='{fnol_id}' found in "
            f"'{path}'. Register file may be out of sync with what "
            f"register_fnol() wrote — check FNOL_REGISTER_PATH."
        )


# ══════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════

@traceable(name="run_triage_agent", run_type="chain")
def run_triage_agent(intake_state: dict, fnol_id: str,
                      fnol_received_at: str = None,
                      progress_callback=None) -> FNOLTriageState:
    """
    Run the FNOL Triage Agent.

    Args:
        intake_state    : FNOLIntakeState dict from run_intake_agent()
        fnol_id         : Registered FNOL ID
        fnol_received_at: Timestamp string (defaults to now)
        progress_callback: Optional callable(step_label: str) -> None,
            called at each major step boundary so the caller can show
            which step is currently running. Unlike the intake agent's
            equivalent, this can be passed straight through as a normal
            parameter — run_triage_agent() is a plain function, not a
            LangGraph node, so there's no state-schema stripping risk to
            work around. Never required — a callback that raises is
            swallowed internally so a UI bug can't break triage
            generation.

    Returns:
        FNOLTriageState with all sections populated.
    """
    if fnol_received_at is None:
        fnol_received_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    fnol_fields = intake_state.get("fnol_fields", {})

    # ── Per-run timing capture (same pattern as fnol_intake_agent.py) ────
    _start_time = time.perf_counter()
    _step_marks: list[tuple[str, float]] = []

    def _report(step_label: str) -> None:
        _step_marks.append((step_label, time.perf_counter()))
        if progress_callback:
            try:
                progress_callback(step_label)
            except Exception:
                pass

    state: FNOLTriageState = {
        "fnol_id":                   fnol_id,
        "triage_generated_at":       datetime.now().isoformat(),
        "adjuster_id":               fnol_fields.get("adjuster_id") or "",
        "claim_snapshot":            {},
        "aging_risk":                {},
        "coverage_result":           [],
        "coverage_applicable":       False,
        "coverage_confidence_overall": "LOW",
        "exclusion_result":          [],
        "waiting_period":            {},
        "claims_history":            {},
        "fraud_signals":             {},
        "doc_checklist":             [],
        "escalation":                {},
        "triage_status":             "IN_PROGRESS",  # replaced by actual status at end
        "escalation_flag":           False,
        "escalation_reason":         None,
        "triage_json_path":          None,
        "triage_pdf_path":           None,
        "timing_report":             {},
        "timing_report_path":        None,
        "agent_status":              "PENDING",
        "error_message":             None,
    }

    try:
        _report("Looking up your policy details…")
        # Step 1: Claim snapshot + policy lookup
        snapshot, policy_row = build_claim_snapshot(
            fnol_fields, fnol_id, fnol_received_at
        )
        state["claim_snapshot"] = snapshot

        _report("Analysing coverage applicability…")
        # Step 2: Coverage analysis
        coverage, applicable, confidence = analyse_coverage(
            fnol_fields, policy_row, snapshot
        )
        state["coverage_result"]             = coverage
        state["coverage_applicable"]         = applicable
        state["coverage_confidence_overall"] = confidence

        _report("Checking policy exclusions…")
        # Step 3: Exclusion analysis
        state["exclusion_result"] = analyse_exclusions(fnol_fields, policy_row)

        _report("Checking waiting period…")
        # Step 4: Waiting period
        state["waiting_period"] = check_waiting_period(fnol_fields, policy_row)

        _report("Reviewing claims history…")
        # Step 5: Claims history
        state["claims_history"] = check_claims_history(fnol_fields, policy_row)

        _report("Evaluating fraud signals…")
        # Step 5.5: Fraud signal detection — reuses waiting_period and
        # claims_history already computed above, rather than re-deriving
        # anything from Excel. Separate module (fraud_signal_agent.py),
        # same call pattern as fnol_triage_pdf.render_triage_pdf() below.
        from fraud_signal_agent import evaluate_fraud_signals
        state["fraud_signals"] = evaluate_fraud_signals(
            fnol_fields         = fnol_fields,
            policy_row          = policy_row,
            extracted_summary   = intake_state.get("extracted_summary", {}),
            waiting_period      = state["waiting_period"],
            claims_history      = state["claims_history"],
            consistency_checks  = intake_state.get("consistency_checks", []),
            pii_store           = intake_state.get("pii_store", {}),
            fnol_received_at    = fnol_received_at,
        )

        _report("Checking document checklist…")
        # Step 6: Document checklist (from intake state)
        checklist, critical_missing = build_doc_checklist(intake_state)
        state["doc_checklist"] = checklist

        _report("Evaluating escalation conditions…")
        # Step 7: Escalation — now also takes fraud_signals as input
        escalation = evaluate_escalation(
            fnol_fields, policy_row,
            state["waiting_period"],
            state["claims_history"],
            state["fraud_signals"],
            fnol_id, fnol_received_at
        )
        state["escalation"]       = escalation
        state["escalation_flag"]  = escalation["escalation_flag"]
        state["escalation_reason"]= escalation["escalation_reason"]

        _report("Calculating aging risk…")
        # Step 8: Aging risk
        state["aging_risk"] = compute_aging_risk(fnol_received_at)

        # Step 9: Triage status
        state["triage_status"] = determine_triage_status(
            policy_row, escalation, critical_missing, state["waiting_period"]
        )

        _report("Saving triage report…")
        # Step 10: Save JSON and PDF
        import os
        from pathlib import Path
        out_dir = Path(_triage_output_path()) / fnol_id
        out_dir.mkdir(parents=True, exist_ok=True)

        json_path = out_dir / f"{fnol_id}_triage.json"
        with open(json_path, "w") as f:
            json.dump(dict(state), f, indent=2, default=str)
        state["triage_json_path"] = str(json_path)

        _report("Generating triage summary PDF…")
        # PDF
        from fnol_triage_pdf import render_triage_pdf
        pdf_path = str(out_dir / f"{fnol_id}_triage.pdf")
        render_triage_pdf(state, pdf_path)
        state["triage_pdf_path"] = pdf_path

        _report("Updating FNOL register…")
        # Step 11: Update FNOL register
        update_fnol_register(
            fnol_id        = fnol_id,
            triage_status  = state["triage_status"],
            escalation_flag= state["escalation_flag"],
            escalation_reason = state["escalation_reason"],
            triage_json_path  = str(json_path),
            triage_pdf_path   = pdf_path,
        )

        # ── Build + save the per-run timing report ──────────────────────
        # Justification: same pattern as fnol_intake_agent.py — real
        # measured durations for THIS run, not estimates. duration_s for
        # each step = time until the NEXT step's _report() call fired;
        # the last step's duration runs until now (right after
        # update_fnol_register returns).
        _end_time = time.perf_counter()
        timing_steps = []
        for idx, (label, ts) in enumerate(_step_marks):
            next_ts = _step_marks[idx + 1][1] if idx + 1 < len(_step_marks) else _end_time
            timing_steps.append({
                "step":           label,
                "duration_s":     round(next_ts - ts, 3),
                "start_offset_s": round(ts - _start_time, 3),
            })
        timing_report = {
            "total_duration_s": round(_end_time - _start_time, 3),
            "steps": timing_steps,
        }
        state["timing_report"] = timing_report

        # Saved alongside the triage JSON/PDF for this fnol_id. Wrapped in
        # try/except: a save failure here must never fail triage generation.
        try:
            timing_report_path = out_dir / f"{fnol_id}_timing_report.json"
            with open(timing_report_path, "w") as f:
                json.dump(timing_report, f, indent=2)
            state["timing_report_path"] = str(timing_report_path)
        except Exception as e:
            logger.warning(f"Timing report save failed: {e}")
            state["timing_report_path"] = None

        state["agent_status"] = "COMPLETE"

        # ── Re-write the triage JSON with FINAL values ──────────────────
        # Justification: the earlier json.dump() (a few lines up) necessarily
        # ran BEFORE triage_pdf_path, timing_report, timing_report_path, and
        # agent_status were set — those all get their real values only
        # after that point in the code. Without this second write, the
        # saved file on disk is a mid-run snapshot showing agent_status:
        # "PENDING", triage_pdf_path: null, and timing_report: {} even
        # though the run genuinely completed — confirmed by a real
        # production file. The in-memory `state` returned to main.py was
        # always correct; only the saved copy was stale. Wrapped in
        # try/except like the timing report save above: a failure here
        # must never fail triage generation, since generation itself has
        # already fully succeeded by this point.
        try:
            with open(json_path, "w") as f:
                json.dump(dict(state), f, indent=2, default=str)
        except Exception as e:
            logger.warning(f"Final triage JSON re-write failed: {e}")

    except Exception as e:
        state["agent_status"]  = "ERROR"
        state["error_message"] = str(e)
        raise

    return state
